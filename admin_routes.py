from flask import Blueprint, render_template, request, jsonify, session, redirect, url_for, flash,send_file
import os
import sqlite3
from werkzeug.utils import secure_filename
from openpyxl.styles import Font
import openpyxl
from datetime import datetime

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_db():
    conn = sqlite3.connect('exam_system.db')
    conn.row_factory = sqlite3.Row
    return conn


@admin_bp.route('/create-exam', methods=['GET', 'POST'])
def create_exam():
    if session.get('role') != 'admin':
        return redirect(url_for('admin_login'))

    if request.method == 'POST':
        title = request.form.get('title')
        description = request.form.get('description')
        duration = request.form.get('duration_minutes')
        passing_score = request.form.get('passing_score')

        conn = get_db()
        cursor = conn.execute(
            'INSERT INTO exams (title, description, duration_minutes, passing_score) VALUES (?, ?, ?, ?)',
            (title, description, duration, passing_score)
        )
        exam_id = cursor.lastrowid
        conn.commit()
        conn.close()

        flash('Exam created successfully!', 'success')
        return redirect(url_for('admin.upload_questions', exam_id=exam_id))

    return render_template('admin/create_exam.html')



@admin_bp.route('/upload-questions/<int:exam_id>', methods=['GET', 'POST'])
def upload_questions(exam_id):
    if session.get('role') != 'admin':
        return redirect(url_for('admin_login'))

    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file uploaded', 'error')
            return redirect(request.url)

        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(request.url)

        if not allowed_file(file.filename):
            flash('Only Excel files (.xlsx, .xls) are allowed!', 'error')
            return redirect(request.url)

        filename = secure_filename(file.filename)
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)

        try:
            wb = openpyxl.load_workbook(filepath)
            sheet = wb.active
            questions = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                question = {
                    'question_text': str(row[0]),
                    'option_a': str(row[1]),
                    'option_b': str(row[2]),
                    'option_c': str(row[3]),
                    'option_d': str(row[4]),
                    'correct_answer': str(row[5]).upper() if row[5] else 'A'
                }
                questions.append(question)

            conn = get_db()
            for q in questions:
                conn.execute('''INSERT INTO questions 
                                (exam_id, question_text, option_a, option_b, option_c, option_d, correct_answer)
                                VALUES (?, ?, ?, ?, ?, ?, ?)''',
                             (exam_id, q['question_text'], q['option_a'], q['option_b'],
                              q['option_c'], q['option_d'], q['correct_answer']))
            conn.commit()
            conn.close()

            os.remove(filepath)
            flash(f'{len(questions)} questions successfully uploaded!', 'success')
            return redirect(url_for('admin.view_results', exam_id=exam_id))

        except Exception as e:
            flash(f'Error processing file: {str(e)}', 'error')
            if os.path.exists(filepath):
                os.remove(filepath)
            return redirect(request.url)

    conn = get_db()
    exam = conn.execute('SELECT * FROM exams WHERE id = ?', (exam_id,)).fetchone()
    conn.close()

    return render_template('admin/upload_questions.html', exam=exam)



@admin_bp.route('/view-results/<int:exam_id>')
def view_results(exam_id):
    if session.get('role') != 'admin':
        return redirect(url_for('admin_login'))

    conn = get_db()
    exam = conn.execute('SELECT * FROM exams WHERE id = ?', (exam_id,)).fetchone()
    results = conn.execute('''SELECT sa.*, u.username, u.full_name 
                              FROM student_attempts sa 
                              JOIN users u ON sa.student_id = u.id 
                              WHERE sa.exam_id = ? AND sa.status IN ("completed", "terminated")
                              ORDER BY sa.submitted_at DESC''',
                           (exam_id,)).fetchall()
    conn.close()

    return render_template('admin/view_results.html', exam=exam, results=results)


@admin_bp.route('/view-logs/<int:attempt_id>')
def view_logs(attempt_id):
    if session.get('role') != 'admin':
        return redirect(url_for('admin_login'))

    conn = get_db()

    attempt = conn.execute('''
        SELECT sa.*, u.username, u.full_name, e.title 
        FROM student_attempts sa
        JOIN users u ON sa.student_id = u.id
        JOIN exams e ON sa.exam_id = e.id
        WHERE sa.id = ?
    ''', (attempt_id,)).fetchone()


    logs = conn.execute('''
        SELECT * FROM monitoring_logs
        WHERE attempt_id = ?
        ORDER BY timestamp ASC
    ''', (attempt_id,)).fetchall()

    report = conn.execute('''
        SELECT * FROM exam_reports
        WHERE attempt_id = ?
        ORDER BY timestamp DESC LIMIT 1
    ''', (attempt_id,)).fetchone()

    conn.close()

    start_time = end_time = None
    total_seconds = 0
    total_minutes = 0
    warning_logs = []

    if logs:
        logs = [dict(log) for log in logs]

        def parse_time(ts):
            try:
                if 'T' in ts:
                    return datetime.fromisoformat(ts)
                return datetime.strptime(ts, '%Y-%m-%d %H:%M:%S')
            except Exception:
                return None

        start_time = parse_time(logs[0]['timestamp'])
        end_time = parse_time(logs[-1]['timestamp'])

        if start_time and end_time:
            total_seconds = round((end_time - start_time).total_seconds())
            total_minutes = round(total_seconds / 60, 1)

        warning_logs = [log for log in logs if log.get('warning_issued') == 1]

    return render_template(
        'admin/view_logs.html',
        attempt=attempt,
        logs=logs,
        warning_logs=warning_logs,
        report=report,
        total_seconds=total_seconds,
        total_minutes=total_minutes,
        start_time=start_time,
        end_time=end_time
    )

@admin_bp.route('/delete-exam/<int:exam_id>')
def delete_exam(exam_id):
    if session.get('role') != 'admin':
        return redirect(url_for('admin_login'))

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM questions WHERE exam_id = ?', (exam_id,))
    cursor.execute('DELETE FROM monitoring_logs WHERE attempt_id IN (SELECT id FROM student_attempts WHERE exam_id = ?)', (exam_id,))
    cursor.execute('DELETE FROM exam_reports WHERE attempt_id IN (SELECT id FROM student_attempts WHERE exam_id = ?)', (exam_id,))
    cursor.execute('DELETE FROM student_attempts WHERE exam_id = ?', (exam_id,))
    cursor.execute('DELETE FROM exams WHERE id = ?', (exam_id,))

    conn.commit()
    conn.close()

    flash('Exam deleted successfully.', 'success')
    return redirect(url_for('admin_dashboard.html'))


@admin_bp.route('/export-monitoring/<int:attempt_id>')
def export_monitoring(attempt_id):
    conn = sqlite3.connect('exam_system.db')
    conn.row_factory = sqlite3.Row

    logs = conn.execute(
        "SELECT * FROM monitoring_logs WHERE attempt_id = ? ORDER BY id ASC",
        (attempt_id,)
    ).fetchall()
    conn.close()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Monitoring Logs"

    headers = [
        "ID", "Attempt ID", "Event Type", "Face Detected",
        "Gaze Direction", "Head Pose", "Warning Issued",
        "Details", "Created At"
    ]

    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for log in logs:
        sheet.append([
            log["id"],
            log["attempt_id"],
            log["event_type"],
            log["face_detected"],
            log["gaze_direction"],
            log["head_pose"],
            log["warning_issued"],
            log["details"],
            log["created_at"]
        ])

    filename = f"/tmp/monitoring_attempt_{attempt_id}.xlsx"
    workbook.save(filename)

    return send_file(filename, as_attachment=True)